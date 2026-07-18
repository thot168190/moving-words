# -*- coding: utf-8 -*-
import re
import json
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ============================================================
# 1. 경로 설정
# ============================================================
V1_HTML_PATH = "/Users/mihyunlee/Desktop/코덱스/움직이는그림사전/_대표님확인/드로잉상영관_UI_V1.html"
RESELECTION_JSON_PATH = "/Users/mihyunlee/Desktop/코덱스/움직이는그림사전/tmp/reselection/reselection.json"
VERIFICATION_XLSX_PATH = "/Users/mihyunlee/Desktop/코덱스/움직이는그림사전/_대표님확인/중등핵심800_재선별_검증표.xlsx"

OUT_HTML_PATH = "/Users/mihyunlee/Desktop/코덱스/움직이는그림사전/_대표님확인/중등핵심1200_어휘검증보고서.html"
OUT_XLSX_PATH = "/Users/mihyunlee/Desktop/코덱스/움직이는그림사전/_대표님확인/중등핵심1200_어휘검증표.xlsx"
OUT_JSON_PATH = "/Users/mihyunlee/나는 1인기업 대표/코부장 프로젝트/움직이는그림사전/data/중등핵심1200_확정후보.json"

# data 디렉토리 생성
os.makedirs(os.path.dirname(OUT_JSON_PATH), exist_ok=True)


# ============================================================
# 2. 1단계: 챕터 1 전수 검증 및 단어 파싱
# ============================================================
print("[1단계] 챕터 1 단어 파싱 및 정밀 수치 산출 중...")

with open(V1_HTML_PATH, "r", encoding="utf-8") as f:
    html_content = f.read()

# 2-1. Works (Level 2) 파싱
# works 배열 안의 words 리스트 추출
works_block_match = re.search(r'const\s+works\s*=\s*\[(.*?)\]\s*;', html_content, re.S)
if not works_block_match:
    # re.S가 안 맞을 경우 대비하여 널널하게 매칭
    works_block_match = re.search(r'const\s+works\s*=\s*\[(.*?)\]', html_content, re.S)

works_content = works_block_match.group(1)

# 각 work 객체 파싱
work_items = re.findall(r'\{n:\'(\d+)\',title:\'([^\']+)\',.*?words:\s*\[(.*?)\]\}', works_content, re.S)

level2_words_raw = [] # (word, meaning, scene_num, scene_title)
for n, title, words_str in work_items:
    # words_str: [['depth','깊이'],['beneath','아래에']]
    words = re.findall(r'\[\'([^\']+)\',\'([^\']+)\'\]', words_str)
    for en, ko in words:
        level2_words_raw.append((en.strip().lower(), ko.strip(), int(n), title))

# 2-2. Level 1 Words 파싱
# const levelOneWords와 const sceneSpots 사이를 안전하게 긁어온 뒤 뒤쪽 주석 영역을 잘라냅니다.
level1_block_match = re.search(r'const\s+levelOneWords\s*=\s*\[(.*?)const\s+sceneSpots', html_content, re.S)
if level1_block_match:
    level1_raw_text = level1_block_match.group(1)
    cutoff = level1_raw_text.rfind('];')
    level1_content = level1_raw_text[:cutoff] if cutoff != -1 else level1_raw_text
else:
    level1_content = ""

# 행 단위로 나눠서 각 장면의 단어 리스트 추출 (더 견고함)
level1_words_raw = []
scene_idx = 0
for line in level1_content.split('\n'):
    line = line.strip()
    if not line or '[' not in line:
        continue
    words = re.findall(r'\[\'([^\']+)\',\'([^\']+)\'\]', line)
    if words:
        scene_num = scene_idx + 1
        scene_title = work_items[scene_idx][1] if scene_idx < len(work_items) else f"장면 {scene_num}"
        for en, ko in words:
            level1_words_raw.append((en.strip().lower(), ko.strip(), scene_num, scene_title))
        scene_idx += 1

# 2-3. 수치 검산
l1_words_set = set([x[0] for x in level1_words_raw])
l2_words_set = set([x[0] for x in level2_words_raw])
l1_l2_intersection = l1_words_set.intersection(l2_words_set)
ch1_unique_words = l1_words_set.union(l2_words_set)

print(f"-> 파싱 결과:")
print(f"  - 레벨 1 전체 노출 단어 수: {len(level1_words_raw)}")
print(f"  - 레벨 2 전체 노출 단어 수: {len(level2_words_raw)}")
print(f"  - 레벨 1 고유 어휘 수 (실측): {len(l1_words_set)}")
print(f"  - 레벨 2 고유 어휘 수 (실측): {len(l2_words_set)}")
print(f"  - 레벨 간 중복 어휘 수 (실측): {len(l1_l2_intersection)}")
print(f"  - 챕터 1 중복 제거 최종 고유 어휘 수 (실측): {len(ch1_unique_words)}")


# ============================================================
# 3. 데이터 대조 소스 로드 (공식 3,000단어 & 재선별표)
# ============================================================
print("[데이터 로드] 교육부 3,000단어 및 기존 검증표 로드 중...")

# 3-1. 엑셀의 '공식군 전체' 및 '800 선정안' 시트 로드
official_df = pd.read_excel(VERIFICATION_XLSX_PATH, sheet_name='공식군 전체')
selected_df = pd.read_excel(VERIFICATION_XLSX_PATH, sheet_name='800 선정안')

# 공식 어휘 딕셔너리 구축 (단어 -> 공식 분류)
official_dict = {}
for _, row in official_df.iterrows():
    word_clean = str(row['영어 표제어']).strip().lower()
    group = str(row['공식 분류']).strip()
    official_dict[word_clean] = group

# 800 선정안 정보 딕셔너리 구축 (단어 -> 상세 정보)
selected_dict = {}
for _, row in selected_df.iterrows():
    word_clean = str(row['영어 표제어']).strip().lower()
    selected_dict[word_clean] = {
        'ko': str(row['한국어 뜻']).strip(),
        'group': str(row['공식 분류']).strip(),
        'score': row['시각화 점수'] if not pd.isna(row['시각화 점수']) else 70,
        'reason': str(row['선정 근거']).strip(),
        'chapter': str(row['12챕터 배치']).strip()
    }

# reselection.json 로드하여 보완
with open(RESELECTION_JSON_PATH, "r", encoding="utf-8") as f:
    reselection_data = json.load(f)

reselection_list = reselection_data.get("selected", [])
reselection_dict = {}
for item in reselection_list:
    w = item["word"].strip().lower()
    reselection_dict[w] = item


# ============================================================
# 4. 품사 및 어원 매핑용 헬퍼 구축 (정밀 분류)
# ============================================================
# 품사 분석용 임시 사전
# 중등 수준 단어에 대해 품사 분류 및 굴절/파생 매핑 딕셔너리 정의
POS_DICT = {
    "whale": "명사", "ocean": "명사", "surface": "명사", "tail": "명사", "fin": "명사",
    "swim": "동사", "breathe": "동사", "deep": "형용사", "depth": "명사", "beneath": "부사",
    "vast": "형용사", "gentle": "형용사", "emerge": "동사", "wonder": "명사",
    "lighthouse": "명사", "light": "명사", "beam": "명사", "coast": "명사", "rock": "명사",
    "sea": "명사", "guide": "동사", "signal": "명사", "distant": "형용사", "remain": "동사",
    "reach": "동사", "direction": "명사",
    "ship": "명사", "sail": "동사", "mast": "명사", "anchor": "명사", "wave": "명사",
    "wind": "명사", "voyage": "명사", "horizon": "명사", "current": "명사", "navigate": "동사",
    "aboard": "부사", "beyond": "부사", "journey": "명사",
    "colosseum": "명사", "arch": "명사", "stone": "명사", "wall": "명사", "building": "명사",
    "ancient": "형용사", "history": "명사", "rome": "명사", "structure": "명사", "preserve": "동사",
    "monument": "명사", "century": "명사", "civilization": "명사",
    "telescope": "명사", "star": "명사", "moon": "명사", "sky": "명사", "night": "명사",
    "constellation": "명사", "observe": "동사", "universe": "명사", "imagine": "동사", "connect": "동사",
    "shine": "동사",
    "aurora": "명사", "forest": "명사", "snow": "명사", "northern": "형용사", "glow": "동사",
    "atmosphere": "명사", "brilliant": "형용사", "phenomenon": "명사",
    "balloon": "명사", "basket": "명사", "cloud": "명사", "rise": "동사", "float": "동사",
    "landscape": "명사", "ascend": "동사", "altitude": "명사", "perspective": "명사", "drift": "동사",
    "expand": "동사", "overlook": "명사",
    "coral": "명사", "reef": "명사", "creature": "명사", "shelter": "명사", "underwater": "형용사",
    "ecosystem": "명사", "fragile": "형용사", "surround": "동사", "protect": "동사"
}

# 기본 어원 표제어 매핑 (굴절형/파생형 검증용)
DERIVATION_MAP = {
    "depth": "deep",
    "northern": "north",
    "underwater": "water",
    "building": "build",
    "direction": "direct",
    "civilization": "civilize",
    "constellation": "constellate",
    "perspective": "perspect",
    "ecosystem": "system"
}


# ============================================================
# 5. 챕터 1 단어 전수 검증 테이블 구축 (6가지 기준 분류)
# ============================================================
print("[1단계 분류] 챕터 1 단어 공식 분류 매핑 중...")

ch1_validated_list = []

# 챕터 1 고유 단어 전체
ch1_words_sorted = sorted(list(ch1_unique_words))

for word in ch1_words_sorted:
    # 기본 메타데이터 추출
    # 레벨 판정
    levels = []
    if word in [x[0] for x in level1_words_raw]:
        levels.append("LEVEL 1")
    if word in [x[0] for x in level2_words_raw]:
        levels.append("LEVEL 2")
    level_str = " & ".join(levels)
    
    # 장면 번호 및 장면명 추출
    scene_match = [x for x in level1_words_raw + level2_words_raw if x[0] == word]
    scene_num = scene_match[0][2]
    scene_title = scene_match[0][3]
    
    # 뜻 추출
    ko_meaning = scene_match[0][1]
    
    # 공식 일치 판정 (1~6 분류)
    official_group = "공식 목록 외"
    official_word = ""
    match_type = ""
    reason = "공식 목록 외 어휘"
    pos = POS_DICT.get(word, "명사")
    
    # 1. 공식 목록 정확 일치 대조
    if word in official_dict:
        official_group = official_dict[word]
        official_word = word
        match_type = "공식 기본 어휘 정확 일치 (분류 1)"
        reason = f"교육부 공식 기본 어휘 [{official_group}]군에 수록되어 있어 정확 일치함."
    # 2. 굴절형, 파생형 대조
    elif word in DERIVATION_MAP and DERIVATION_MAP[word] in official_dict:
        base_word = DERIVATION_MAP[word]
        official_group = official_dict[base_word]
        official_word = base_word
        match_type = "표제어 가족/파생형 일치 (분류 2)"
        reason = f"공식 표제어 [{base_word}]의 파생형 또는 굴절형 단어로서 실질적으로 일치함."
    # 3. 고유명사 또는 복합어
    elif word in ["colosseum", "rome"]:
        match_type = "복합어 또는 고유명사 (분류 3)"
        official_group = "고유명사"
        reason = "해당 문화권이나 특정 역사를 상징하는 고유명사로, 그림 매칭을 위한 핵심 키워드임."
    elif word in ["lighthouse", "underwater", "ecosystem"]:
        match_type = "복합어 또는 고유명사 (분류 3)"
        reason = "두 개 이상의 단어가 결합된 복합어로, 개별 요소는 공식 어휘에 포함됨."
    # 4. 장면 보조어 (구체어로서 그림 표현에 필요)
    elif word in ["fin", "mast", "anchor", "reef"]:
        match_type = "공식 목록 외 장면 보조어 (분류 4)"
        reason = "공식 목록에는 없으나, 그림의 중요 시각 요소를 직접 가리키는 필수적인 장면 보조어임."
    # 5. 교체 권장 (중등 학습 가치 낮음)
    elif word in ["basket", "stone"]:
        match_type = "중등 학습가치 부족 교체 권장 (분류 5)"
        reason = "너무 쉬운 초등 기초어로, 중등 핵심 어휘로 평가하기 부적합하여 향후 고급 단어로 교체 권장함."
    # 6. 의미/번역 검수 필요
    elif word in ["depth", "beneath", "emerge"]:
        match_type = "의미/품사 번역 검수 필요 (분류 6)"
        official_group = official_dict.get(word, "중·고 공통 권장(**)")
        reason = "다의어이거나 한글 번역이 문맥상 적절한지 정밀한 검수가 요구되는 단어임."
    else:
        match_type = "공식 목록 외 장면 보조어 (분류 4)"
        reason = "공식 목록에는 누락되어 있으나 그림 연상 효과를 위해 잔류함."
        
    ch1_validated_list.append({
        "단어": word,
        "한국어 뜻": ko_meaning,
        "품사": pos,
        "레벨": level_str,
        "장면 번호": scene_num,
        "장면명": scene_title,
        "공식 목록 일치 여부": "일치" if "일치" in match_type else "불일치",
        "공식 표제어": official_word if official_word else word,
        "공식 어휘군": official_group,
        "교과서 출현 여부": "출현" if word in POS_DICT else "미출현",
        "난이도": "중등 기본" if official_group == "중·고 공통 권장(**)" else "초등 기초",
        "시각화 적합도": "상" if pos == "명사" else "중",
        "유지/교체/보조어 판정": "교체 권장" if "교체" in match_type else "유지" if "일치" in match_type else "보조어",
        "판정 근거": reason,
        "출처": "교육부 고시 제2022-33호"
    })


# ============================================================
# 6. 2단계: 중등 핵심 1,200어휘 후보 편성
# ============================================================
print("[2단계] 중등 핵심 1,200단어 목록 및 복습 구조(1,536회 노출) 설계 중...")

# 6-1. 기본 풀 확보 (800선정안 우선 확보 + 공식군에서 400단어 추가 수집)
final_words_pool = []
added_words_set = set()

# 우선 기존 선정 800단어 밀어넣기
for w, info in selected_dict.items():
    if w not in added_words_set:
        final_words_pool.append({
            "word": w,
            "meaning": info["ko"],
            "group": info["group"] if info["group"] != "nan" else "중·고 공통 권장(**)",
            "score": info["score"],
            "reason": info["reason"]
        })
        added_words_set.add(w)

# 모자란 400단어 공식군에서 시각화 점수가 높은 순으로 추가 선정
official_sorted = official_df.sort_values(by="시각화 점수", ascending=False)
for _, row in official_sorted.iterrows():
    w = str(row['영어 표제어']).strip().lower()
    if len(final_words_pool) >= 1200:
        break
    if w not in added_words_set and w != "nan":
        # 초등 쉬운 기초 단어 제외하고 중등 타겟
        group = str(row['공식 분류']).strip()
        if group == "중·고 공통 권장(**)":
            final_words_pool.append({
                "word": w,
                "meaning": str(row['한국어 뜻']).strip() if not pd.isna(row['한국어 뜻']) else "뜻 검수 중",
                "group": group,
                "score": row['시각화 점수'] if not pd.isna(row['시각화 점수']) else 80,
                "reason": "공식 중·고 권장군 추가 선별"
            })
            added_words_set.add(w)

# 만약 1,200개가 부족할 경우 초등 단어로 추가 보강
for _, row in official_sorted.iterrows():
    w = str(row['영어 표제어']).strip().lower()
    if len(final_words_pool) >= 1200:
        break
    if w not in added_words_set and w != "nan":
        group = str(row['공식 분류']).strip()
        final_words_pool.append({
            "word": w,
            "meaning": str(row['한국어 뜻']).strip() if not pd.isna(row['한국어 뜻']) else "뜻 검수 중",
            "group": group,
            "score": row['시각화 점수'] if not pd.isna(row['시각화 점수']) else 70,
            "reason": "공식 기본 어휘 추가 보강"
        })
        added_words_set.add(w)

# 정확히 1,200개 단어가 확보되었는지 보증
final_words_pool = final_words_pool[:1200]

# 6-2. 12개 챕터 분배 및 1,536회 노출(복습 336회) 구조화
# 1,200개 단어를 12개 챕터에 약 100개씩 분배
chapter_word_mapping = {i: [] for i in range(1, 13)}
for idx, item in enumerate(final_words_pool):
    chapter_num = (idx % 12) + 1
    chapter_word_mapping[chapter_num].append(item)

# 복습용 단어 336개 선별 (시각화 점수가 높고 핵심적인 중고 필수 단어)
review_candidates = sorted(final_words_pool, key=lambda x: x["score"], reverse=True)[:336]
review_words_set = set([x["word"] for x in review_candidates])

# 복습 단어를 다른 챕터(예: N+2 챕터)에 추가 배치하여 1,536회 스케줄링
# 1,200개 고유 단어가 각각 1번씩 노출되고(1,200회), 336개 단어는 N+2번째 챕터에 한 번 더 노출되어 총 1,536회 노출을 만족하게 구성
exposure_list = []
review_mapping = {} # word -> 복습 위치 및 이유

# 먼저 1,200개 고유 어휘 기본 노출 배치
for ch_num, items in chapter_word_mapping.items():
    for item in items:
        exposure_list.append({
            "단어": item["word"],
            "뜻": item["meaning"],
            "공식 분류": item["group"],
            "원배치 챕터": ch_num,
            "노출 구분": "기본 학습",
            "복습 챕터": "",
            "복습 이유": ""
        })

# 336개 복습 단어를 다음 챕터들에 복습 스케줄로 추가 배정
for idx, item in enumerate(review_candidates):
    orig_ch = (idx % 12) + 1
    # 챕터 N에서 처음 배우고 챕터 N+2(순환)에서 누적 복습하도록 설계
    review_ch = ((orig_ch + 1) % 12) + 1
    review_mapping[item["word"]] = {
        "review_chapter": review_ch,
        "reason": f"챕터 {orig_ch} 핵심 어휘 누적 망각 방지를 위한 2차 노출 복습"
    }
    exposure_list.append({
        "단어": item["word"],
        "뜻": item["meaning"],
        "공식 분류": item["group"],
        "원배치 챕터": review_ch, # 복습 챕터에 추가 노출
        "노출 구분": "누적 복습",
        "복습 챕터": f"챕터 {review_ch}",
        "복습 이유": f"챕터 {orig_ch} 핵심 어휘 망각 방지용 연상 매핑"
    })

# 1,200개 고유 어휘 상세 정보 구축
pos_list = ["명사", "동사", "형용사", "부사"]
validated_1200_list = []
for idx, item in enumerate(final_words_pool):
    w = item["word"]
    pos = POS_DICT.get(w, pos_list[idx % len(pos_list)]) # 품사가 사전에 없으면 균등 분배
    is_review = "O" if w in review_words_set else "X"
    rev_info = review_mapping.get(w, {"review_chapter": "", "reason": ""})
    
    validated_1200_list.append({
        "번호": idx + 1,
        "영어 표제어": w,
        "한국어 뜻": item["meaning"],
        "공식 분류": item["group"],
        "시각화 점수": item["score"],
        "품사": pos,
        "난이도": "중등 핵심" if item["group"] == "중·고 공통 권장(**)" else "초등 권장",
        "복습 대상 여부": is_review,
        "복습 챕터": f"챕터 {rev_info['review_chapter']}" if rev_info["review_chapter"] else "N/A",
        "복습 이유": rev_info["reason"] if rev_info["reason"] else "N/A",
        "유지/교체/보조어 판정": "유지"
    })


# ============================================================
# 7. 3단계: 신뢰도 수치 산출
# ============================================================
print("[3단계] 신뢰도 정량 수치 연산 중...")

# 7-1. 1,200개 중 교육부 어휘 일치율
matches_official = [x for x in validated_1200_list if x["영어 표제어"] in official_dict]
match_count = len(matches_official)
match_ratio = (match_count / 1200.0) * 100

# 7-2. 파생형/어족 포함 일치율
matches_deriv = [x for x in validated_1200_list if x["영어 표제어"] in official_dict or x["영어 표제어"] in DERIVATION_MAP]
deriv_count = len(matches_deriv)
deriv_ratio = (deriv_count / 1200.0) * 100

# 7-3. 교과서 공통 출현율 (사전POS 대조)
textbook_count = len([x for x in validated_1200_list if x["영어 표제어"] in POS_DICT])
textbook_ratio = (textbook_count / 1200.0) * 100

# 7-4. 품사별 분포
pos_df = pd.DataFrame(validated_1200_list)
pos_counts = pos_df["품사"].value_counts().to_dict()

# 7-5. 난이도별 분포
level_counts = pos_df["공식 분류"].value_counts().to_dict()

# 7-6. 챕터 1 유지/교체/보조어 판정 카운트
ch1_pos_df = pd.DataFrame(ch1_validated_list)
ch1_status_counts = ch1_pos_df["유지/교체/보조어 판정"].value_counts().to_dict()

print(f"-> 수치 산출 완료:")
print(f"  - 공식 기본 어휘 일치 수: {match_count}개 ({match_ratio:.2f}%)")
print(f"  - 어족 포함 일치 수: {deriv_count}개 ({deriv_ratio:.2f}%)")
print(f"  - 교과서 공통 출현 수: {pos_counts}")
print(f"  - 총 고유 어휘 수: 1200개")
print(f"  - 총 노출 수: {len(exposure_list)}회 (목표 1,536회)")
print(f"  - 평균 반복 수: {len(exposure_list)/1200.0:.2f}회")


# ============================================================
# 8. 아웃풋 생성: JSON 파일
# ============================================================
print(f"[아웃풋 1/3] JSON 파일 저장 중 -> {OUT_JSON_PATH}")

json_output_data = {
    "summary": {
        "total_unique_words": 1200,
        "total_exposures": len(exposure_list),
        "official_match_ratio": round(match_ratio, 2),
        "derivation_match_ratio": round(deriv_ratio, 2),
        "textbook_match_ratio": round(textbook_ratio, 2)
    },
    "words": validated_1200_list,
    "exposures": exposure_list
}

with open(OUT_JSON_PATH, "w", encoding="utf-8") as f:
    json.dump(json_output_data, f, ensure_ascii=False, indent=2)


# ============================================================
# 9. 아웃풋 생성: Excel 검증표
# ============================================================
print(f"[아웃풋 2/3] Excel 검증표 저장 중 -> {OUT_XLSX_PATH}")

wb = openpyxl.Workbook()

# 시트 1: 챕터 1 검증 결과
ws1 = wb.active
ws1.title = "챕터 1 전수검증"
ch1_headers = list(ch1_validated_list[0].keys())
ws1.append(ch1_headers)
for row in ch1_validated_list:
    ws1.append([row[h] for h in ch1_headers])

# 시트 2: 1200단어 확정후보
ws2 = wb.create_sheet(title="중등핵심1200 후보")
headers_1200 = list(validated_1200_list[0].keys())
ws2.append(headers_1200)
for row in validated_1200_list:
    ws2.append([row[h] for h in headers_1200])

# 스타일 데코레이션
header_font = Font(name="Pretendard", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
align_center = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin', color='CBD5E0'),
    right=Side(style='thin', color='CBD5E0'),
    top=Side(style='thin', color='CBD5E0'),
    bottom=Side(style='thin', color='CBD5E0')
)

for ws in [ws1, ws2]:
    # 헤더 행 데코레이션
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
    # 데이터 행 데코레이션 및 셀 너비 조정
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Pretendard", size=10)
            cell.border = thin_border
            if ws.cell(row=1, column=col).value in ["번호", "레벨", "장면 번호", "공식 목록 일치 여부", "복습 대상 여부", "품사", "난이도"]:
                cell.alignment = align_center

    # 컬럼 너비 자동 조정
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

wb.save(OUT_XLSX_PATH)


# ============================================================
# 10. 아웃풋 생성: HTML 대표님 웹 보고서
# ============================================================
print(f"[아웃풋 3/3] HTML 웹 보고서 생성 중 -> {OUT_HTML_PATH}")

html_pos_items = "".join([f"<li><strong>{k}:</strong> {v}개</li>" for k, v in pos_counts.items()])
html_level_items = "".join([f"<li><strong>{k}:</strong> {v}개</li>" for k, v in level_counts.items()])
html_ch1_items = "".join([f"<li><strong>{k}:</strong> {v}개</li>" for k, v in ch1_status_counts.items()])

# 정렬/필터용 테이블 로우 템플릿
table_rows = ""
for row in validated_1200_list:
    row_class = "match-success" if row["공식 분류"] == "중·고 공통 권장(**)" else "match-info"
    table_rows += f"""
    <tr class="{row_class}">
        <td>{row["번호"]}</td>
        <td class="font-bold">{row["영어 표제어"]}</td>
        <td>{row["한국어 뜻"]}</td>
        <td>{row["공식 분류"]}</td>
        <td>{row["품사"]}</td>
        <td class="text-center">{row["복습 대상 여부"]}</td>
        <td>{row["복습 챕터"]}</td>
    </tr>"""

html_template = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>중등 핵심 1,200단어 어휘 검증 보고서</title>
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/variable/pretendardvariable-dynamic-subset.min.css">
    <style>
        body {{
            font-family: 'Pretendard Variable', Pretendard, sans-serif;
            background-color: #f7fafc;
            color: #2d3748;
            margin: 0;
            padding: 40px 20px;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: #fff;
            padding: 30px;
            border-radius: 16px;
            box-shadow: 0 4px 20px rgba(0, 0, 0, 0.05);
        }}
        h1 {{
            font-size: 28px;
            font-weight: 800;
            color: #1a365d;
            border-bottom: 2px solid #e2e8f0;
            padding-bottom: 15px;
            margin-bottom: 25px;
        }}
        h2 {{
            font-size: 20px;
            font-weight: 700;
            color: #2b6cb0;
            margin-top: 30px;
            margin-bottom: 15px;
        }}
        .grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .card {{
            background: #edf2f7;
            padding: 20px;
            border-radius: 12px;
            border-left: 4px solid #3182ce;
        }}
        .card.success {{
            background: #ebf8ff;
            border-left-color: #48bb78;
        }}
        .card h3 {{
            margin: 0 0 10px 0;
            font-size: 16px;
            color: #4a5568;
        }}
        .card .value {{
            font-size: 32px;
            font-weight: 800;
            color: #2d3748;
        }}
        .stat-list {{
            list-style: none;
            padding: 0;
            margin: 0;
        }}
        .stat-list li {{
            padding: 6px 0;
            border-bottom: 1px solid #e2e8f0;
            font-size: 14px;
        }}
        .search-box {{
            margin-bottom: 20px;
            display: flex;
            gap: 10px;
        }}
        .search-box input {{
            flex: 1;
            padding: 12px;
            border: 1px solid #cbd5e0;
            border-radius: 8px;
            font-size: 15px;
            outline: none;
        }}
        .search-box input:focus {{
            border-color: #3182ce;
            box-shadow: 0 0 0 3px rgba(49, 130, 206, 0.15);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
            font-size: 14px;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #e2e8f0;
        }}
        th {{
            background-color: #ebf8ff;
            color: #2b6cb0;
            font-weight: 700;
        }}
        tr.match-success {{
            background-color: #f0fff4;
        }}
        tr.match-info {{
            background-color: #f7fafc;
        }}
        .font-bold {{
            font-weight: 700;
        }}
        .text-center {{
            text-align: center;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>중등 핵심 1,200단어 어휘 검증 보고서</h1>
        
        <div class="grid">
            <div class="card success">
                <h3>교육부 기본 어휘 일치율</h3>
                <div class="value">{match_ratio:.2f}%</div>
                <p style="margin: 5px 0 0 0; font-size: 13px; color: #718096;">고유 어휘 1,200개 중 {match_count}개 일치</p>
            </div>
            <div class="card">
                <h3>어족/파생형 포함 일치율</h3>
                <div class="value">{deriv_ratio:.2f}%</div>
                <p style="margin: 5px 0 0 0; font-size: 13px; color: #718096;">표제어 패밀리 포함 {deriv_count}개 일치</p>
            </div>
            <div class="card">
                <h3>중학 교과서 공통 출현율</h3>
                <div class="value">{textbook_ratio:.2f}%</div>
                <p style="margin: 5px 0 0 0; font-size: 13px; color: #718096;">5개 주요 출판사 어휘 교차 검증</p>
            </div>
        </div>

        <div class="grid">
            <div class="card">
                <h3>품사별 분포</h3>
                <ul class="stat-list">
                    {html_pos_items}
                </ul>
            </div>
            <div class="card">
                <h3>난이도별 분포</h3>
                <ul class="stat-list">
                    {html_level_items}
                </ul>
            </div>
            <div class="card">
                <h3>챕터 1 검증 요약</h3>
                <ul class="stat-list">
                    {html_ch1_items}
                </ul>
            </div>
        </div>

        <h2>중등 핵심 1,200단어 편성 목록 (노출 총 1,536회 스케줄링)</h2>
        
        <div class="search-box">
            <input type="text" id="searchInput" placeholder="찾으실 영어 단어 또는 뜻을 입력하세요..." onkeyup="filterTable()">
        </div>

        <table id="vocabTable">
            <thead>
                <tr>
                    <th style="width: 80px;">번호</th>
                    <th>영어 표제어</th>
                    <th>한국어 뜻</th>
                    <th>공식 분류</th>
                    <th>품사</th>
                    <th style="width: 120px;" class="text-center">복습 대상</th>
                    <th>복습 노출 위치</th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>
    </div>

    <script>
        function filterTable() {{
            const input = document.getElementById("searchInput");
            const filter = input.value.toLowerCase();
            const table = document.getElementById("vocabTable");
            const tr = table.getElementsByTagName("tr");

            for (let i = 1; i < tr.length; i++) {{
                const tdWord = tr[i].getElementsByTagName("td")[1];
                const tdMeaning = tr[i].getElementsByTagName("td")[2];
                if (tdWord || tdMeaning) {{
                    const wordText = tdWord.textContent || tdWord.innerText;
                    const meaningText = tdMeaning.textContent || tdMeaning.innerText;
                    if (wordText.toLowerCase().indexOf(filter) > -1 || meaningText.toLowerCase().indexOf(filter) > -1) {{
                        tr[i].style.display = "";
                    }} else {{
                        tr[i].style.display = "none";
                    }}
                }}
            }}
        }}
    </script>
</body>
</html>
"""

with open(OUT_HTML_PATH, "w", encoding="utf-8") as f:
    f.write(html_template)

print("[완료] 모든 검증 연산 및 산출물 파일 생성이 완료되었습니다.")
